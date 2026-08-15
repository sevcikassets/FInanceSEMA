from __future__ import annotations

import argparse
import hashlib
import unicodedata
import uuid
from collections import Counter
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete, select, update
from sqlalchemy.orm import Session

from .db import Base, SessionLocal, engine
from .models import (
    Asset,
    AssetCost,
    AssetType,
    CostCategory,
    DailyStatistic,
    ExchangeRate,
    ImportBatch,
    LoanMovement,
    Party,
    Portfolio,
    PortfolioPosition,
    StockTransaction,
    TickerDescription,
    WatchlistStock,
)


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def as_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value).replace("\xa0", "").replace(" ", "").replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def as_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def checksum(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def get_or_create_party(db: Session, name: str | None, kind: str = "unknown") -> Party | None:
    if not name:
        return None
    existing = db.scalar(select(Party).where(Party.name == name))
    if existing:
        if existing.kind == "unknown" and kind != "unknown":
            existing.kind = kind
        return existing
    party = Party(name=name, kind=kind)
    db.add(party)
    db.flush()
    return party


def get_or_create_cost_category(db: Session, portfolio_id: uuid.UUID, name: str | None) -> CostCategory | None:
    """Same get-or-create pattern as get_or_create_party above, scoped to
    one Subjekt's category dictionary instead of the global Party table -
    used both by the Excel import and the manual add/edit-cost endpoint in
    main.py, so a category typed by hand lands in the same dictionary a
    later import would create it in."""
    if not name:
        return None
    existing = db.scalar(select(CostCategory).where(CostCategory.portfolio_id == portfolio_id, CostCategory.name == name))
    if existing:
        return existing
    category = CostCategory(portfolio_id=portfolio_id, name=name)
    db.add(category)
    db.flush()
    return category


def get_or_create_asset_type(
    db: Session, portfolio_id: uuid.UUID, name: str | None, calculation_mode: str = "none"
) -> AssetType | None:
    """Same get-or-create pattern as get_or_create_cost_category. calculation_mode
    only applies when the type doesn't exist yet - an existing type's mode is
    never overwritten by import/migration, since an admin may have deliberately
    customized it via the "Typy majetku" agenda."""
    if not name:
        return None
    existing = db.scalar(select(AssetType).where(AssetType.portfolio_id == portfolio_id, AssetType.name == name))
    if existing:
        return existing
    asset_type = AssetType(portfolio_id=portfolio_id, name=name, calculation_mode=calculation_mode, required_fields=[])
    db.add(asset_type)
    db.flush()
    return asset_type


def normalize_match_text(value: str | None) -> str:
    if not value:
        return ""
    text = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return " ".join(part for part in text.lower().replace("-", " ").split() if part not in {"byt"})


def reset_imported_tables(db: Session, portfolio_id: uuid.UUID) -> None:
    # ExchangeRate/TickerDescription are intentionally NOT included here - they
    # are shared/global reference data (CZK rates, ticker->name lookups), not
    # one Subjekt's property. Wiping them on every import would delete another
    # Subjekt's contributed rate/ticker history (see import_exchange_rates,
    # which upserts instead of blind-inserting for the same reason).
    # Asset.linked_asset_id is self-referential (a Hypotéka points at the
    # property it finances) - null it out first so the bulk delete below
    # isn't deleting rows that still reference each other.
    db.execute(update(Asset).where(Asset.portfolio_id == portfolio_id).values(linked_asset_id=None))
    for model in [
        DailyStatistic,
        PortfolioPosition,
        StockTransaction,
        WatchlistStock,
        AssetCost,
        Asset,
        LoanMovement,
    ]:
        db.execute(delete(model).where(model.portfolio_id == portfolio_id))


def import_loans(db: Session, wb, portfolio_id: uuid.UUID) -> int:
    if "Půjčky Pohyby" not in wb.sheetnames:
        return 0
    ws = wb["Půjčky Pohyby"]
    count = 0
    for r in range(2, ws.max_row + 1):
        raw_date = ws.cell(r, 1).value
        amount = as_decimal(ws.cell(r, 4).value)
        lender_name = as_text(ws.cell(r, 2).value)
        borrower_name = as_text(ws.cell(r, 3).value)
        movement_date = as_date(raw_date)
        # The sheet has monthly/yearly subtotal rows baked right in between the
        # real movements (column A holds a text label like "Leden 2023" or
        # "2023" instead of a date, and lender/borrower are blank) - these are
        # not real loan movements and must not be imported as if they were.
        # The app builds its own (collapsible) month/year subtotals instead.
        if movement_date is None:
            continue
        if amount is None:
            continue
        movement = LoanMovement(
            portfolio_id=portfolio_id,
            movement_date=movement_date,
            lender=get_or_create_party(db, lender_name),
            borrower=get_or_create_party(db, borrower_name),
            amount=amount,
            interest_rate=as_decimal(ws.cell(r, 5).value),
            interest_period=as_text(ws.cell(r, 6).value),
            planned_end_date=as_date(ws.cell(r, 7).value),
            completed_at=as_date(ws.cell(r, 8).value),
            description=as_text(ws.cell(r, 9).value),
            source_row=r,
        )
        db.add(movement)
        count += 1
    return count


def import_assets(db: Session, wb, portfolio_id: uuid.UUID) -> int:
    if "Investice" not in wb.sheetnames:
        return 0
    ws = wb["Investice"]
    count = 0
    year_cols = [(c, str(ws.cell(3, c).value)) for c in range(16, ws.max_column + 1) if ws.cell(3, c).value]
    for r in range(4, ws.max_row + 1):
        code = as_text(ws.cell(r, 2).value)
        name = as_text(ws.cell(r, 5).value)
        if not code or not name:
            continue
        plan = {
            year: float(value)
            for c, year in year_cols
            if (value := as_decimal(ws.cell(r, c).value)) is not None
        }
        asset_type_name = as_text(ws.cell(r, 4).value)
        asset = Asset(
            portfolio_id=portfolio_id,
            code=code,
            owner=get_or_create_party(db, as_text(ws.cell(r, 3).value), "owner"),
            asset_type=asset_type_name,
            asset_type_id=get_or_create_asset_type(db, portfolio_id, asset_type_name).id if asset_type_name else None,
            name=name,
            total_value=as_decimal(ws.cell(r, 6).value),
            own_funds=as_decimal(ws.cell(r, 7).value),
            borrowed_amount=as_decimal(ws.cell(r, 8).value),
            lender_name=as_text(ws.cell(r, 9).value),
            borrowed_from=as_date(ws.cell(r, 10).value),
            borrowed_to=as_date(ws.cell(r, 11).value),
            interest_rate=as_decimal(ws.cell(r, 12).value),
            loan_years=as_decimal(ws.cell(r, 13).value),
            fixed_until=as_date(ws.cell(r, 14).value),
            payment=as_decimal(ws.cell(r, 15).value),
            annual_interest_plan=plan,
            source_row=r,
        )
        db.add(asset)
        count += 1
    return count


def move_zapujcka_assets_to_loans(db: Session, portfolio_id: uuid.UUID) -> int:
    """"Zápůjčka" (money lent OUT, earning interest) belongs in Půjčky
    (LoanMovement already has lender_id/borrower_id/interest_rate/
    planned_end_date) - not in Majetek as an asset type. Moves every asset
    still carrying the legacy free-text asset_type == "Zápůjčka" into a real
    LoanMovement record and deletes the Asset row. Operates on the legacy
    text column, before any asset_type_id backfill, so no orphaned
    "Zápůjčka" AssetType ever gets created. Self-idempotent: once moved, no
    matching Asset rows remain for a rerun to find. Also called at the end
    of import_workbooks() so a future Excel re-import doesn't reintroduce
    these rows until the next app restart."""
    candidates = db.scalars(select(Asset).where(Asset.portfolio_id == portfolio_id, Asset.asset_type == "Zápůjčka")).all()
    count = 0
    for asset in candidates:
        db.add(
            LoanMovement(
                portfolio_id=portfolio_id,
                movement_date=asset.borrowed_from,
                lender_id=asset.owner_id,
                borrower=get_or_create_party(db, asset.name, "borrower"),
                amount=asset.own_funds,
                interest_rate=asset.interest_rate,
                planned_end_date=asset.borrowed_to,
                description=f"Zápůjčka (převedeno z majetku {asset.code})",
            )
        )
        db.delete(asset)
        count += 1
    return count


def split_debt_assets_into_linked_liability(db: Session, portfolio_id: uuid.UUID) -> int:
    """A mortgage is a debt on a property, not a property attribute - splits
    an Asset row that has both valuation fields (total_value/own_funds) AND
    full loan terms into two rows: the original keeps just the valuation,
    and a new linked "Hypotéka" Asset (calculation_mode="debt_interest")
    carries the loan fields, pointing back via linked_asset_id.

    Candidate rows must have borrowed_amount set and nonzero (excludes a
    zero-debt property with no real loan) AND interest_rate/borrowed_from set
    (excludes a mortgage whose terms were never filled in - same "missing
    inputs" gate project_annual_interest already uses, and it must stay
    calculation_mode="none" since there's nothing to project) AND
    source_row set (only ever splits Excel-imported rows - a manually
    created Asset via the CRUD form that deliberately keeps loan fields
    inline must never get silently split apart on the next restart).

    Self-idempotent: once split, the original's borrowed_amount becomes
    NULL, so it no longer matches the candidate query on a rerun. Also
    called at the end of import_workbooks() for the same reason as
    move_zapujcka_assets_to_loans above.
    """
    loan_fields = ("borrowed_amount", "lender_name", "borrowed_from", "borrowed_to", "interest_rate", "loan_years", "fixed_until", "payment")
    candidates = db.scalars(
        select(Asset).where(
            Asset.portfolio_id == portfolio_id,
            Asset.borrowed_amount.isnot(None),
            Asset.borrowed_amount != 0,
            Asset.interest_rate.isnot(None),
            Asset.borrowed_from.isnot(None),
            Asset.source_row.isnot(None),
        )
    ).all()
    count = 0
    for original in candidates:
        hypoteka_type = get_or_create_asset_type(db, portfolio_id, "Hypotéka", calculation_mode="debt_interest")
        linked = Asset(
            portfolio_id=portfolio_id,
            code=f"{original.code}-HYP",
            name=f"Hypotéka {original.name}",
            owner_id=original.owner_id,
            asset_type=hypoteka_type.name,
            asset_type_id=hypoteka_type.id,
            linked_asset_id=original.id,
            **{field: getattr(original, field) for field in loan_fields},
        )
        db.add(linked)
        for field in loan_fields:
            setattr(original, field, None)
        count += 1
    return count


def find_asset_by_name(db: Session, portfolio_id: uuid.UUID, sheet_name: str) -> Asset | None:
    sheet_tokens = normalize_match_text(sheet_name).split()
    for asset in db.scalars(select(Asset).where(Asset.portfolio_id == portfolio_id)).all():
        asset_text = normalize_match_text(asset.name)
        if sheet_tokens and all(token in asset_text for token in sheet_tokens):
            return asset
    return None


def find_header_row(ws, required: set[str], max_scan_rows: int = 10) -> tuple[int, list[str | None]] | None:
    for row_number in range(1, min(ws.max_row, max_scan_rows) + 1):
        headers = [as_text(ws.cell(row_number, c).value) for c in range(1, ws.max_column + 1)]
        if required.issubset({header for header in headers if header}):
            return row_number, headers
    return None


def import_asset_cost_sheets(db: Session, wb, portfolio_id: uuid.UUID) -> int:
    count = 0
    skip = {
        "Rekaptiulace ÚVĚRY",
        "Půjčky Pohyby",
        "Investice",
        "Akcie",
        "Sledované akcie",
        "Akcie statistika",
        "Graf",
        "Portfolio",
        "List1",
        "List2",
        "Analyza",
        "Kurzy",
        "Akcie popis",
    }
    for ws in wb.worksheets:
        if ws.title in skip:
            continue
        if ws.max_row < 2:
            continue
        header_info = find_header_row(ws, {"Datum", "Částka"})
        if header_info is None:
            continue
        header_row, headers = header_info
        asset = find_asset_by_name(db, portfolio_id, ws.title)
        for r in range(header_row + 1, ws.max_row + 1):
            amount = as_decimal(ws.cell(r, headers.index("Částka") + 1).value)
            item = as_text(ws.cell(r, headers.index("Účel") + 1).value) if "Účel" in headers else None
            if amount is None and not item:
                continue
            db.add(
                AssetCost(
                    portfolio_id=portfolio_id,
                    asset=asset,
                    cost_date=as_date(ws.cell(r, headers.index("Datum") + 1).value),
                    supplier=as_text(ws.cell(r, headers.index("Firma") + 1).value) if "Firma" in headers else None,
                    item=item or "Náklad",
                    amount=amount,
                    source_sheet=ws.title,
                    source_row=r,
                )
            )
            count += 1
    return count


def import_property_costs(db: Session, path: Path, portfolio_id: uuid.UUID) -> int:
    wb = load_workbook(path, data_only=True, read_only=True)
    if "Finance" not in wb.sheetnames:
        return 0
    asset = db.scalar(select(Asset).where(Asset.portfolio_id == portfolio_id, Asset.code == "RD-KVASICE"))
    if asset is None:
        nemovitost_type = get_or_create_asset_type(db, portfolio_id, "Nemovitost")
        asset = Asset(
            portfolio_id=portfolio_id,
            code="RD-KVASICE",
            name="RD Kvasice",
            asset_type=nemovitost_type.name,
            asset_type_id=nemovitost_type.id,
        )
        db.add(asset)
        db.flush()
    ws = wb["Finance"]
    count = 0
    for r in range(6, ws.max_row + 1):
        raw_date = ws.cell(r, 1).value
        item = as_text(ws.cell(r, 4).value)
        amount = as_decimal(ws.cell(r, 6).value)
        is_monthly_summary = raw_date is not None and as_date(raw_date) is None and item is None
        if is_monthly_summary:
            continue
        if not item and amount is None:
            continue
        category = get_or_create_cost_category(db, portfolio_id, as_text(ws.cell(r, 5).value))
        db.add(
            AssetCost(
                portfolio_id=portfolio_id,
                asset=asset,
                cost_date=as_date(raw_date),
                payer=get_or_create_party(db, as_text(ws.cell(r, 2).value), "payer"),
                supplier=as_text(ws.cell(r, 3).value),
                item=item or "Náklad",
                # New imports go straight to the dictionary-backed
                # category_id, not the legacy free-text `category` column -
                # see CostCategory/AssetCost.category_id in models.py.
                category_id=category.id if category else None,
                amount=amount,
                note=as_text(ws.cell(r, 7).value),
                source_sheet="Finance RD Kvasice.xlsx/Finance",
                source_row=r,
            )
        )
        count += 1
    return count


def import_stocks(db: Session, wb, portfolio_id: uuid.UUID) -> int:
    if "Akcie" not in wb.sheetnames:
        return 0
    ws = wb["Akcie"]
    count = 0
    for r in range(4, ws.max_row + 1):
        ticker = as_text(ws.cell(r, 8).value)
        name = as_text(ws.cell(r, 6).value)
        movement = as_text(ws.cell(r, 4).value)
        if not ticker and not name and not movement:
            continue
        db.add(
            StockTransaction(
                portfolio_id=portfolio_id,
                traded_on=as_date(ws.cell(r, 2).value),
                instrument_type=as_text(ws.cell(r, 3).value),
                movement_type=movement,
                cluster=as_text(ws.cell(r, 5).value),
                instrument_name=name,
                isin=as_text(ws.cell(r, 7).value),
                ticker=ticker,
                market=as_text(ws.cell(r, 9).value),
                quantity=as_decimal(ws.cell(r, 10).value),
                unit_price_ccy=as_decimal(ws.cell(r, 11).value),
                limit_ai=as_decimal(ws.cell(r, 12).value),
                gross_amount_ccy=as_decimal(ws.cell(r, 13).value),
                currency=as_text(ws.cell(r, 14).value),
                fee_ccy=as_decimal(ws.cell(r, 15).value),
                fee_czk=as_decimal(ws.cell(r, 16).value),
                amount_czk=as_decimal(ws.cell(r, 17).value),
                current_price=as_decimal(ws.cell(r, 18).value),
                difference_czk=as_decimal(ws.cell(r, 19).value),
                difference_pct=as_decimal(ws.cell(r, 20).value),
                description=as_text(ws.cell(r, 21).value),
                source_row=r,
            )
        )
        count += 1
    return count


def import_watchlist_stocks(db: Session, wb, portfolio_id: uuid.UUID) -> int:
    if "Sledované akcie" not in wb.sheetnames:
        return 0
    ws = wb["Sledované akcie"]
    count = 0
    for r in range(5, ws.max_row + 1):
        ticker = as_text(ws.cell(r, 7).value)
        name = as_text(ws.cell(r, 5).value)
        reason = as_text(ws.cell(r, 3).value)
        if not ticker and not name and not reason:
            continue
        db.add(
            WatchlistStock(
                portfolio_id=portfolio_id,
                watched_on=as_date(ws.cell(r, 2).value),
                reason=reason,
                quantity=as_decimal(ws.cell(r, 4).value),
                name=name,
                isin=as_text(ws.cell(r, 6).value),
                ticker=ticker,
                limit_price=as_decimal(ws.cell(r, 8).value),
                current_price=as_decimal(ws.cell(r, 9).value),
                currency=as_text(ws.cell(r, 10).value),
                difference_pct=as_decimal(ws.cell(r, 11).value),
                week_52_max=as_decimal(ws.cell(r, 12).value),
                week_52_state_pct=as_decimal(ws.cell(r, 13).value),
                note=as_text(ws.cell(r, 14).value),
                owned_value=as_decimal(ws.cell(r, 15).value),
                profit_loss_pct=as_decimal(ws.cell(r, 16).value),
                new_price=as_decimal(ws.cell(r, 17).value),
                source_row=r,
            )
        )
        count += 1
    return count


def import_exchange_rates(db: Session, wb) -> int:
    # Global/shared reference data (see reset_imported_tables) - no longer
    # wiped before every import, so this must upsert by (rate_date, currency)
    # rather than blind-insert, or a second Subjekt's import would hit that
    # unique constraint on any date/currency the first import already covered.
    if "Kurzy" not in wb.sheetnames:
        return 0
    ws = wb["Kurzy"]
    count = 0
    for r in range(2, ws.max_row + 1):
        rate_date = as_date(ws.cell(r, 3).value)
        currency = as_text(ws.cell(r, 4).value)
        rate = as_decimal(ws.cell(r, 5).value)
        if not rate_date or not currency or rate is None:
            continue
        existing = db.scalar(select(ExchangeRate).where(ExchangeRate.rate_date == rate_date, ExchangeRate.currency == currency))
        if existing is None:
            db.add(ExchangeRate(rate_date=rate_date, currency=currency, rate_to_czk=rate))
        else:
            existing.rate_to_czk = rate
        count += 1
    return count


def import_ticker_descriptions(db: Session, wb) -> int:
    if "Akcie popis" not in wb.sheetnames:
        return 0
    ws = wb["Akcie popis"]
    count = 0
    for r in range(2, ws.max_row + 1):
        ticker = as_text(ws.cell(r, 2).value)
        if not ticker:
            continue
        db.merge(
            TickerDescription(
                ticker=ticker,
                name=as_text(ws.cell(r, 3).value),
                isin=as_text(ws.cell(r, 4).value),
                description=as_text(ws.cell(r, 5).value),
            )
        )
        count += 1
    return count


def import_portfolio(db: Session, wb, portfolio_id: uuid.UUID) -> int:
    if "Portfolio" not in wb.sheetnames:
        return 0
    ws = wb["Portfolio"]
    count = 0
    for r in range(3, ws.max_row + 1):
        ticker = as_text(ws.cell(r, 1).value)
        if not ticker:
            continue
        db.merge(
            PortfolioPosition(
                portfolio_id=portfolio_id,
                ticker=ticker,
                name=as_text(ws.cell(r, 2).value),
                quantity=as_decimal(ws.cell(r, 3).value),
                current_price=as_decimal(ws.cell(r, 4).value),
                currency=as_text(ws.cell(r, 5).value),
                market_value_czk=as_decimal(ws.cell(r, 6).value),
                invested_czk=as_decimal(ws.cell(r, 7).value),
                profit_czk=as_decimal(ws.cell(r, 8).value),
                profit_pct=as_decimal(ws.cell(r, 9).value),
                portfolio_share_pct=as_decimal(ws.cell(r, 10).value),
                first_buy_date=as_date(ws.cell(r, 11).value),
                source_row=r,
            )
        )
        count += 1
    return count


def import_daily_statistics(db: Session, wb, portfolio_id: uuid.UUID) -> int:
    if "Akcie statistika" not in wb.sheetnames:
        return 0
    ws = wb["Akcie statistika"]
    count = 0
    for r in range(6, ws.max_row + 1):
        stat_date = as_date(ws.cell(r, 2).value)
        if not stat_date:
            continue
        db.merge(
            DailyStatistic(
                portfolio_id=portfolio_id,
                stat_date=stat_date,
                bought_eur=as_decimal(ws.cell(r, 3).value),
                total_eur=as_decimal(ws.cell(r, 4).value),
                eur_in_czk=as_decimal(ws.cell(r, 5).value),
                value_eur=as_decimal(ws.cell(r, 6).value),
                eur_rate=as_decimal(ws.cell(r, 7).value),
                bought_usd=as_decimal(ws.cell(r, 9).value),
                total_usd=as_decimal(ws.cell(r, 10).value),
                usd_in_czk=as_decimal(ws.cell(r, 11).value),
                value_usd=as_decimal(ws.cell(r, 12).value),
                usd_rate=as_decimal(ws.cell(r, 13).value),
                bought_czk=as_decimal(ws.cell(r, 15).value),
                total_czk=as_decimal(ws.cell(r, 16).value),
                value_czk=as_decimal(ws.cell(r, 17).value),
                invested_czk=as_decimal(ws.cell(r, 19).value),
                total_value_czk=as_decimal(ws.cell(r, 20).value),
                unrealized_profit_czk=as_decimal(ws.cell(r, 21).value),
                dividends=as_decimal(ws.cell(r, 23).value),
                dividends_total=as_decimal(ws.cell(r, 24).value),
                profit_pct=as_decimal(ws.cell(r, 26).value),
                daily_profit_czk=as_decimal(ws.cell(r, 27).value),
                alerts=as_text(ws.cell(r, 28).value),
            )
        )
        count += 1
    return count


def import_workbooks(finance_path: Path, property_costs_path: Path | None = None, *, portfolio_id: uuid.UUID) -> dict[str, int]:
    Base.metadata.create_all(bind=engine)
    counts: Counter[str] = Counter()
    with SessionLocal() as db:
        batch = ImportBatch(source_file=str(finance_path), source_checksum=checksum(finance_path))
        db.add(batch)
        db.flush()
        try:
            reset_imported_tables(db, portfolio_id)
            wb = load_workbook(finance_path, data_only=True, read_only=False, keep_vba=True)
            counts["loans"] = import_loans(db, wb, portfolio_id)
            counts["assets"] = import_assets(db, wb, portfolio_id)
            # SessionLocal is autoflush=False (see db.py) - the two functions
            # below query Asset via `select(...)`, which would not otherwise
            # see import_assets()'s just-added rows yet.
            db.flush()
            counts["loans"] += move_zapujcka_assets_to_loans(db, portfolio_id)
            counts["assets_split"] = split_debt_assets_into_linked_liability(db, portfolio_id)
            counts["asset_costs"] += import_asset_cost_sheets(db, wb, portfolio_id)
            counts["stock_transactions"] = import_stocks(db, wb, portfolio_id)
            counts["watchlist_stocks"] = import_watchlist_stocks(db, wb, portfolio_id)
            counts["exchange_rates"] = import_exchange_rates(db, wb)
            counts["ticker_descriptions"] = import_ticker_descriptions(db, wb)
            counts["portfolio_positions"] = import_portfolio(db, wb, portfolio_id)
            counts["daily_statistics"] = import_daily_statistics(db, wb, portfolio_id)
            if property_costs_path:
                counts["asset_costs"] += import_property_costs(db, property_costs_path, portfolio_id)
            batch.status = "done"
            batch.row_counts = dict(counts)
            batch.finished_at = datetime.now(UTC)
            db.commit()
        except Exception:
            db.rollback()
            batch = db.get(ImportBatch, batch.id)
            batch.status = "failed"
            batch.finished_at = datetime.now(UTC)
            db.commit()
            raise
    return dict(counts)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--finance", required=True, type=Path)
    parser.add_argument("--property-costs", type=Path)
    parser.add_argument("--portfolio-id", type=uuid.UUID, help="Cílový Subjekt (default: první existující, viz /portfolios)")
    args = parser.parse_args()
    Base.metadata.create_all(bind=engine)
    with SessionLocal() as db:
        portfolio_id = args.portfolio_id
        if portfolio_id is None:
            portfolio = db.scalar(select(Portfolio).order_by(Portfolio.created_at))
            if portfolio is None:
                raise SystemExit("Žádný Subjekt v databázi neexistuje - spusťte appku alespoň jednou (vytvoří výchozí) nebo zadejte --portfolio-id.")
            portfolio_id = portfolio.id
    counts = import_workbooks(args.finance, args.property_costs, portfolio_id=portfolio_id)
    for key, value in counts.items():
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()
